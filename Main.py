# the line below imports  libary openpyxl which enables us to  read xl files
import openpyxl as xl

wb =xl.load_workbook('employeedata.xlsx')# the file (database) is been loaded to the script 
sheet=wb['Feuil1']# this line tells the compiler that the active cell is Feuil1 and it should take contents from it
old_email = 'helpinghands.cm'
new_email = 'handsinhands.org'

# Updating the emails in the data base
for i in range (2, sheet.max_row+1):# tells the compiler to iterate over the second element of each column
    cell=sheet.cell(i,2)
    if old_email in cell.value:
        # the line below tells the compiler to acces the element in the email column and replace with the new domain name
        updated_Email=(cell.value).replace(old_email,new_email)
        # attributing the content (data) to be stored in the cell
        sheet.cell(i,2).value = updated_Email
wb.save('employeedata.csv')# this file contains the updated email